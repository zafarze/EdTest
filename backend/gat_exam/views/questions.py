import json
import openpyxl
from rest_framework import viewsets, filters, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.exceptions import PermissionDenied, ValidationError
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from django.http import HttpResponse
from django.db import transaction
from django.db.models import Q
from django_filters.rest_framework import DjangoFilterBackend

# Импорты моделей и сериализаторов
from ..models import Question, Choice, Topic
from ..serializers import QuestionSerializer
from ..permissions import IsQuestionSecurityClearance

# 👇 Импорт утилиты безопасности
from ..utils import get_allowed_school_ids

# 👇 Импорт утилиты для сжатия фото
from ..services.image_optimizer import optimize_image 
# 👇 Импорт AI сервисов
from ..services.ai_service import generate_distractors_ai, analyze_question_ai, parse_file_with_ai

class QuestionViewSet(viewsets.ModelViewSet):
    serializer_class = QuestionSerializer
    # 🔐 Только авторизованные + проверка Clearance (если используется)
    permission_classes = [IsAuthenticated, IsQuestionSecurityClearance]
    
    # 📂 Парсеры для приема файлов и JSON
    parser_classes = (MultiPartParser, FormParser, JSONParser)
    
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['topic', 'question_type', 'difficulty', 'variant']
    search_fields = ['text']
    
    def get_queryset(self):
        """
        🔥 ЦЕНТРАЛЬНАЯ ЗАЩИТА ДАННЫХ
        Пользователь видит вопросы ТОЛЬКО если имеет доступ к их Теме.
        """
        # Жадная загрузка (Join) для скорости
        queryset = Question.objects.select_related('topic', 'topic__subject')\
                                   .prefetch_related('choices')\
                                   .order_by('-id')

        user = self.request.user
        
        # 1. Получаем ID школ, доступных юзеру
        allowed_ids = get_allowed_school_ids(user)

        # 2. Если Admin (None) -> полный доступ
        if allowed_ids is None:
            return queryset

        # 3. Фильтрация:
        # - Вопросы из тем, привязанных к моим школам
        # - Вопросы из тем, которые я создал сам
        return queryset.filter(
            Q(topic__schools__id__in=allowed_ids) | Q(topic__author=user)
        ).distinct()

    # --- 🔥 5. БЫСТРАЯ СМЕНА ПРАВИЛЬНОГО ОТВЕТА ---
    @action(detail=True, methods=['post'], url_path='set-correct')
    def set_correct(self, request, pk=None):
        question = self.get_object() # get_object сам проверит права через get_queryset
        choice_id = request.data.get('choice_id')

        if not choice_id:
            return Response({"error": "ID варианта обязателен"}, status=400)

        with transaction.atomic():
            # Сбрасываем все галочки
            question.choices.update(is_correct=False)
            # Ставим новую
            updated = Choice.objects.filter(id=choice_id, question=question).update(is_correct=True)
            
            if updated == 0:
                return Response({"error": "Вариант не найден или чужой"}, status=404)

        return Response({"status": "success"})

    def create(self, request, *args, **kwargs):
        """Обертка для create с логированием ошибок валидации"""
        serializer = self.get_serializer(data=request.data)
        if not serializer.is_valid():
            print("\n❌ ОШИБКА ВАЛИДАЦИИ (Questions):")
            print(serializer.errors)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        self.perform_create(serializer)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    def perform_create(self, serializer):
        user = self.request.user
        topic = serializer.validated_data.get('topic')
        text = serializer.validated_data.get('text')
        
        # 1. ПРОВЕРКА ДОСТУПА К ТЕМЕ (Security Check)
        allowed_ids = get_allowed_school_ids(user)
        if allowed_ids is not None:
            # Проверка: Тема в моей школе ИЛИ я автор темы
            has_access = topic.schools.filter(id__in=allowed_ids).exists() or topic.author == user
            if not has_access:
                raise PermissionDenied("У вас нет прав добавлять вопросы в эту тему (чужая школа).")

        # 2. ПРОВЕРКА ЭКСПЕРТА
        if hasattr(user, 'profile') and user.profile.role == 'expert':
            if not user.profile.assigned_subjects.filter(id=topic.subject.id).exists():
                raise PermissionDenied(f"Вы не эксперт по предмету '{topic.subject.name}'.")

        # 3. ЗАЩИТА ОТ ДУБЛИКАТОВ (Текст вопроса)
        if Question.objects.filter(topic=topic, text__iexact=text.strip()).exists():
            raise ValidationError({"detail": "Такой вопрос уже существует в этой теме!"})

        # 4. ВАЛИДАЦИЯ ВАРИАНТОВ (Дубликаты ответов)
        choices_json = self.request.data.get('choices_data')
        if choices_json:
            try:
                c_data = json.loads(choices_json)
                texts = [c.get('text', '').strip() for c in c_data if c.get('text', '').strip()]
                if len(texts) != len(set(texts)):
                    raise ValidationError({"detail": "Варианты ответов дублируются!"})
            except: pass

        # 5. СОХРАНЕНИЕ ВОПРОСА
        image = self.request.FILES.get('image')
        optimized_img = optimize_image(image) if image else None
        
        question = serializer.save(image=optimized_img)
        
        # 6. СОХРАНЕНИЕ ВАРИАНТОВ ОТВЕТОВ
        if choices_json:
            try:
                choices_data = json.loads(choices_json)
                for idx, c_data in enumerate(choices_data):
                    # Берем картинку для конкретного варианта (если есть)
                    c_img = self.request.FILES.get(f'choice_image_{idx}')
                    Choice.objects.create(
                        question=question,
                        text=c_data.get('text', ''),
                        is_correct=c_data.get('is_correct', False),
                        image=optimize_image(c_img) if c_img else None
                    )
            except Exception as e:
                print(f"Error saving choices: {e}")

    def perform_update(self, serializer):
        image = self.request.FILES.get('image')
        if image:
            serializer.save(image=optimize_image(image))
        else:
            serializer.save()

    # --- ИМПОРТ EXCEL / AI (ПОЛНАЯ ЛОГИКА) ---
    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser], url_path='import_excel')
    def import_excel(self, request):
        file_obj = request.FILES.get('file')
        topic_id = request.data.get('topic')

        if not file_obj or not topic_id:
            return Response({"error": "Файл и тема обязательны"}, status=400)

        # 1. Проверка прав доступа к теме ПЕРЕД импортом
        try:
            target_topic = Topic.objects.get(id=topic_id)
            
            # Security Check через utils
            allowed_ids = get_allowed_school_ids(request.user)
            if allowed_ids is not None:
                has_access = target_topic.schools.filter(id__in=allowed_ids).exists() or target_topic.author == request.user
                if not has_access:
                    return Response({"error": "Нет доступа к этой теме (чужая школа)"}, status=403)

            # Эксперт check
            if hasattr(request.user, 'profile') and request.user.profile.role == 'expert':
                if not request.user.profile.assigned_subjects.filter(id=target_topic.subject.id).exists():
                    return Response({"error": "Это не ваш профильный предмет"}, status=403)
                    
        except Topic.DoesNotExist:
             return Response({"error": "Тема не найдена"}, status=404)

        filename = file_obj.name.lower()
        created_count = 0
        duplicates_count = 0

        # === ВЕТКА 1: КЛАССИЧЕСКИЙ EXCEL (.xlsx) ===
        if filename.endswith('.xlsx') or filename.endswith('.xls'):
            try:
                wb = openpyxl.load_workbook(file_obj)
                ws = wb.active
                
                with transaction.atomic():
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        # Проверка на пустую строку
                        if not row or not row[0]: continue
                        
                        # Парсинг строки (Текст, Сложность, Тип, A, B, C, D, Правильный)
                        q_text, q_diff, q_type, opt_a, opt_b, opt_c, opt_d, correct_letter = row[:8]
                        q_text_str = str(q_text).strip()

                        # Пропускаем дубликаты
                        if Question.objects.filter(topic_id=topic_id, text__iexact=q_text_str).exists():
                            duplicates_count += 1
                            continue
                        
                        # Создаем вопрос
                        question = Question.objects.create(
                            topic_id=topic_id,
                            text=q_text_str,
                            difficulty=q_diff.lower() if q_diff else 'medium',
                            question_type=q_type.lower() if q_type else 'single'
                        )

                        # Создаем варианты
                        options = [opt_a, opt_b, opt_c, opt_d]
                        letters = ['A', 'B', 'C', 'D']
                        correct_val = str(correct_letter).upper().strip()

                        for i, opt in enumerate(options):
                            if opt: # Если вариант не пустой
                                Choice.objects.create(
                                    question=question,
                                    text=str(opt).strip(),
                                    is_correct=(letters[i] == correct_val)
                                )
                        created_count += 1

                return Response({"status": "success", "processed": created_count, "duplicates": duplicates_count})

            except Exception as e:
                return Response({"error": f"Ошибка Excel: {str(e)}"}, status=500)

        # === ВЕТКА 2: AI IMPORT (DOCX, PDF, IMG) ===
        else:
            print(f"🤖 Запуск AI-импорта для файла: {filename}")
            
            # Вызываем наш AI парсер
            ai_questions = parse_file_with_ai(file_obj, filename)
            
            if not ai_questions:
                return Response({"status": "error", "processed": 0, "message": "AI не смог распознать вопросы."}, status=200)
            
            with transaction.atomic():
                for q_data in ai_questions:
                    q_text = q_data.get('text', '').strip()
                    if not q_text: continue

                    # Проверка дубликатов
                    if Question.objects.filter(topic_id=topic_id, text__iexact=q_text).exists():
                        duplicates_count += 1
                        continue

                    # Создаем вопрос
                    question = Question.objects.create(
                        topic_id=topic_id,
                        text=q_text,
                        difficulty=q_data.get('difficulty', 'medium'),
                        question_type=q_data.get('question_type', 'single')
                    )

                    # Создаем варианты
                    for c_data in q_data.get('choices', []):
                        Choice.objects.create(
                            question=question,
                            text=c_data.get('text', ''),
                            is_correct=c_data.get('is_correct', False)
                        )
                    created_count += 1
            
            return Response({
                "status": "success", 
                "processed": created_count, 
                "duplicates": duplicates_count, 
                "method": "AI"
            })

    # --- СКАЧАТЬ ШАБЛОН ---
    @action(detail=False, methods=['get'])
    def download_template(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Questions"
        ws.append(["Текст вопроса", "Сложность", "Тип", "Вариант A", "Вариант B", "Вариант C", "Вариант D", "Правильный (A-D)"])
        ws.append(["2+2=?", "easy", "single", "3", "4", "5", "6", "B"])
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="questions_template.xlsx"'
        wb.save(response)
        return response

    # --- AI: ГЕНЕРАЦИЯ ДИСТРАКТОРОВ ---
    @action(detail=False, methods=['post'], url_path='ai-distractors')
    def ai_distractors(self, request):
        text = request.data.get('question_text')
        correct = request.data.get('correct_answer')
        if not text or not correct:
            return Response({"error": "Нужен текст вопроса и правильный ответ"}, status=400)
        
        return Response({"distractors": generate_distractors_ai(text, correct)})

    # --- AI: АНАЛИЗ ВОПРОСА ---
    @action(detail=False, methods=['post'], url_path='ai-analyze')
    def ai_analyze(self, request):
        text = request.data.get('text', '')
        choices_raw = request.data.get('choices')
        choices = []
        if choices_raw:
             try:
                 choices = json.loads(choices_raw) if isinstance(choices_raw, str) else choices_raw
             except: pass
                 
        image = request.FILES.get('image')
        return Response(analyze_question_ai(text, choices, image))